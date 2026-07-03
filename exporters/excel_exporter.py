
import openpyxl
import os
from datetime import date
from pathlib import Path
from data.models import Company, HistoricalPrice, IncomeStatement, BalanceSheet, CashflowStatement, FinancialRatios


def export(
    company: Company,
    historical_prices: list[HistoricalPrice],
    income_statements: list[IncomeStatement],
    balance_sheets: list[BalanceSheet],
    cash_flows: list[CashflowStatement],
    ratios: FinancialRatios,
    ai_summary: str | None,
    ticker: str,
    export_dir: Path,
) -> Path:
    """Build a 7-sheet Excel workbook from Pydantic models, save it, and open it."""
    export_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Company Overview"

    historical_ws = wb.create_sheet("Historical Prices")
    income_statement_ws = wb.create_sheet("Income Statement")
    balance_sheet_ws = wb.create_sheet("Balance Sheet")
    cashflow_statement_ws = wb.create_sheet("Cashflow Statement")
    financial_ratios_ws = wb.create_sheet("Financial Ratios")
    ai_summary_ws = wb.create_sheet("AI Summary")

    write_company_overview(ws, company)
    write_historical_prices(historical_ws, historical_prices)
    write_income_statement(income_statement_ws, income_statements)
    write_balance_sheet(balance_sheet_ws, balance_sheets)
    write_cashflow_statement(cashflow_statement_ws, cash_flows)
    write_financial_ratios(financial_ratios_ws, ratios)
    write_ai_summary(ai_summary_ws, ai_summary)

    file_path = export_dir / f"{ticker.upper()}_{date.today().isoformat()}.xlsx"

    wb.save(file_path)

    os.startfile(file_path)

    return file_path


def write_income_statement(ws, income_statements):
    """Write income statement data in transposed layout — metrics as rows, fiscal years as columns."""
    ws["A1"] = "Metric"

    metrics = [
        ("Revenue", "revenue"),
        ("Gross Profit", "gross_profit"),
        ("Operating Income", "operating_income"),
        ("EBIT", "ebit"),
        ("EBITDA", "ebitda"),
        ("Pretax Income", "pretax_income"),
        ("Net Income", "net_income"),
        ("EPS", "eps"),
    ]

    if not income_statements:
        for row_num, (label, attr_name) in enumerate(metrics, start=2):
            ws.cell(row=row_num, column=1).value = label
        return

    for row_num, (label, attr_name) in enumerate(metrics, start=2):
        ws.cell(row=row_num, column=1).value = label

    for col, statement in enumerate(income_statements, start=2):
        ws.cell(row=1, column=col).value = statement.fiscal_year
        for row_num, (label, attr_name) in enumerate(metrics, start=2):
            value = getattr(statement, attr_name, None)
            ws.cell(row=row_num, column=col).value = value


def write_company_overview(ws, company):
    """Write company profile fields as two-column label/value rows."""
    ws["A1"] = "Metric"
    ws["B1"] = "Value"

    metrics = [
    ("Ticker", "ticker"),
    ("Company", "company_name"),
    ("Exchange", "exchange"),
    ("Sector", "sector"),
    ("Industry", "industry"),
    ("Country", "country"),
    ("Employees", "employees"),
    ("Market Cap", "market_cap"),
    ("Enterprise Value", "enterprise_value"),
    ("Current Price", "current_price"),
    ("Currency", "currency"),
    ("Shares Outstanding", "shares_outstanding"),
    ("Beta", "beta"),
    ("Dividend Yield", "dividend_yield"),
    ("52 Week High", "week_52_high"),
    ("52 Week Low", "week_52_low"),
    ]

    for row_num, (label, attr_name) in enumerate(metrics, start=2):
        ws.cell(row=row_num, column=1).value = label
        ws.cell(row=row_num, column=2).value = getattr(company, attr_name, None)
    
def write_historical_prices(ws, historical_prices):
    """Write historical price records in flat layout — one row per date, fields as columns."""
    metrics = [
        ("Date", "date"),
        ("Open", "open"),
        ("High", "high"),
        ("Low", "low"),
        ("Close", "close"),
        ("Volume", "volume"),
        ("Adjusted Close", "adjusted_close"),
    ]

    for col_num, (label, attr_name) in enumerate(metrics, start=1):
        ws.cell(row=1, column=col_num).value = label

    for row, historical_price in enumerate(historical_prices, start=2):
        for col_num, (label, attr_name) in enumerate(metrics, start=1):
            value = getattr(historical_price, attr_name, None)
            ws.cell(row=row, column=col_num).value = value


def write_balance_sheet(ws, balance_sheets):
    """Write balance sheet data in transposed layout — metrics as rows, fiscal years as columns."""
    ws['A1'] = "Metric"

    metrics = [
        ("Cash", "cash"),
        ("Inventory", "inventory"),
        ("Current Assets", "current_assets"),
        ("Total Assets", "total_assets"),
        ("Current Liabilities", "current_liabilities"),
        ("Long Term Debt", "long_term_debt"),
        ("Total Liabilities", "total_liabilities"),
        ("Shareholders Equity", "shareholders_equity"),
    ]

    if not balance_sheets:
        for row_num, (label, attr_name) in enumerate(metrics, start=2):
            ws.cell(row=row_num, column=1).value = label
        return
    
    for row_num, (label, attr_name) in enumerate(metrics, start=2):
        ws.cell(row=row_num, column=1).value = label
    
    for col, balance_sheet in enumerate(balance_sheets, start=2):
        ws.cell(row=1, column=col).value = balance_sheet.fiscal_year
        for row_num, (label, attr_name) in enumerate(metrics, start=2):
            value = getattr(balance_sheet, attr_name, None)
            ws.cell(row=row_num, column=col).value = value

def write_cashflow_statement(ws, cash_flows):
    """Write cash flow data in transposed layout — metrics as rows, fiscal years as columns."""
    ws['A1'] = "Metric"

    metrics = [
        ("Operating Cash Flow", "operating_cash_flow"),
        ("Investing Cash Flow", "investing_cash_flow"),
        ("Financing Cash Flow", "financing_cash_flow"),
        ("Net Cash Change", "net_cash_change"),
        ("Capital Expenditures", "capital_expenditures"),
        ("Free Cash Flow", "free_cash_flow"),
    ]

    if not cash_flows:
        for row_num, (label, attr_name) in enumerate(metrics, start=2):
            ws.cell(row=row_num, column=1).value = label
        return

    for row_num, (label, attr_name) in enumerate(metrics, start=2):
        ws.cell(row=row_num, column=1).value = label
    
    for col, cashflow in enumerate(cash_flows, start=2):
        ws.cell(row=1, column=col).value = cashflow.fiscal_year
        for row_num, (label, attr_name) in enumerate(metrics, start=2):
            value = getattr(cashflow, attr_name, None)
            ws.cell(row=row_num, column=col).value = value

def write_financial_ratios(ws, ratios):
    """Write financial ratio fields as two-column label/value rows."""
    ws["A1"] = "Metric"
    ws["B1"] = "Value"

    metrics = [
        ("PEG Ratio", "peg_ratio"),
        ("PE Ratio", "pe_ratio"),
        ("Forward PE", "forward_pe"),
        ("ROE", "roe"),
        ("ROA", "roa"),
        ("Debt/Equity", "debt_equity"),
        ("Current Ratio", "current_ratio"),
        ("Quick Ratio", "quick_ratio"),
        ("Gross Margin", "gross_margin"),
        ("Operating Margin", "operating_margin"),
        ("Profit Margin", "profit_margin"),
        ("Revenue Growth", "revenue_growth"),
        ("EPS Growth", "eps_growth"),     
    ]

    for row_num, (label, attr_name) in enumerate(metrics, start=2):
        ws.cell(row=row_num, column=1).value = label
        ws.cell(row=row_num, column=2).value = getattr(ratios, attr_name, None)

def write_ai_summary(ws, ai_summary):
    """Write the AI summary string into cell A1, or a fallback message if absent."""
    if not ai_summary:
        ws["A1"] = "AI Summary not available."
    else:
        ws["A1"] = ai_summary
